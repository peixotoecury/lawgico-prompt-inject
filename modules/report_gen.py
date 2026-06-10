"""
Geração de relatório Excel com openpyxl.
"""
from __future__ import annotations
import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY  = '001E36'
NAVY2 = '003B5C'
CYAN  = '00A9E0'
WHITE = 'FFFFFF'
GRAY  = 'F4F8FB'

C_CRITICO = {'bg':'7F1D1D','fg':'FFFFFF'}
C_ALTO    = {'bg':'FEE2E2','fg':'991B1B'}
C_MEDIO   = {'bg':'FEF9C3','fg':'713F12'}
C_LIMPO   = {'bg':'D1FAE5','fg':'065F46'}
C_NIV     = {1:'D1FAE5',2:'A7F3D0',3:'FEF9C3',4:'FEE2E2',5:'FCA5A5'}

def fill(h): return PatternFill('solid', fgColor=h)
def bdb(c='D0E4EE'): return Border(bottom=Side(style='thin', color=c))
def bold(size=10, color=NAVY, name='Segoe UI'): return Font(name=name, size=size, bold=True, color=color)
def reg(size=9, color='17324D', name='Segoe UI'): return Font(name=name, size=size, color=color)
def al(h='left', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _hdr_row(ws, row, cols_labels, bg=NAVY2, fg=WHITE):
    for i, (col, lbl) in enumerate(cols_labels, 1):
        c = ws.cell(row=row, column=i)
        c.value = lbl
        c.font = Font(name='Segoe UI', size=9, bold=True, color=fg)
        c.fill = fill(bg)
        c.alignment = al('center')
        c.border = Border(bottom=Side(style='medium', color=CYAN))


def gerar_excel(resultados: list[dict], cliente: str = "") -> bytes:
    wb = openpyxl.Workbook()

    # ── Aba 1: Resumo ──
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    today = datetime.date.today().strftime('%d/%m/%Y')
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 38
    ws.column_dimensions['H'].width = 3

    # Header
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 48
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 4

    for col in 'ABCDEFGH':
        ws[col+'1'].fill = fill(NAVY)
        ws[col+'2'].fill = fill(NAVY)
        ws[col+'3'].fill = fill(NAVY)
        ws[col+'4'].fill = fill(CYAN)

    ws.merge_cells('B2:G2')
    ws['B2'].value = 'LAWgico Prompt Inject — Relatório de Análise'
    ws['B2'].font = Font(name='Segoe UI', size=14, bold=True, color=WHITE)
    ws['B2'].alignment = al('left', 'center')

    ws.merge_cells('B3:D3')
    ws['B3'].value = f'Peixoto & Cury Advogados  ·  {cliente + "  ·  " if cliente else ""}Gerado em {today}'
    ws['B3'].font = Font(name='Segoe UI', size=9, italic=True, color='90C8E0')
    ws['B3'].alignment = al('left', 'center')

    # KPIs
    total   = len(resultados)
    critico = sum(1 for r in resultados if r.get('risco_final', {}).get('risco') == 'CRÍTICO')
    alto    = sum(1 for r in resultados if r.get('risco_final', {}).get('risco') == 'ALTO')
    revisar = sum(1 for r in resultados if r.get('risco_final', {}).get('requer_revisao'))
    n4_5    = sum(1 for r in resultados if r.get('ia', {}).get('agressividade', {}).get('nivel', 0) >= 4)

    ws.row_dimensions[6].height = 52
    kpis = [
        ('B6', total,   'Documentos\nAnalisados', CYAN,    '001E36'),
        ('C6', critico, 'Injection\nCRÍTICO',    '7F1D1D', WHITE),
        ('D6', alto,    'Injection\nALTO',        'B45309', WHITE),
        ('E6', revisar, 'Fila de\nRevisão',       'A16207', '17324D'),
        ('F6', n4_5,    'Agressividade\nNível 4-5','991B1B', WHITE),
    ]
    for addr, val, lbl, bg, fg in kpis:
        c = ws[addr]
        ws.merge_cells(f'{addr}:{addr}')
        c.value = f'{val}\n{lbl}'
        c.font = Font(name='Segoe UI', size=11, bold=True, color=fg)
        c.fill = fill(bg)
        c.alignment = al('center', 'center', wrap=True)

    ws.row_dimensions[7].height = 8
    ws.row_dimensions[8].height = 26

    _hdr_row(ws, 8, [
        ('B', 'Arquivo'), ('C', 'Injection'), ('D', 'Confiança'),
        ('E', 'Nível Agress.'), ('F', 'Valor Estimado'), ('G', 'Resumo Estratégico')
    ])

    row = 9
    for i, r in enumerate(resultados):
        ws.row_dimensions[row].height = 36
        bg = GRAY if i % 2 == 0 else WHITE
        inj = r.get('risco_final', {})
        ia  = r.get('ia', {})
        agr = ia.get('agressividade', {})
        risco = inj.get('risco', '—')
        niv   = agr.get('nivel', 0)

        clr_r = {'CRÍTICO': C_CRITICO, 'ALTO': C_ALTO, 'MÉDIO': C_MEDIO, 'LIMPO': C_LIMPO}.get(risco, C_LIMPO)
        clr_n = C_NIV.get(niv, WHITE)

        for col in 'ABCDEFGH':
            ws[col+str(row)].fill = fill(bg)

        # Arquivo
        c = ws['B'+str(row)]
        c.value = r.get('filename', '')
        c.font = bold(9)
        c.alignment = al('left', 'center', wrap=True)
        c.border = bdb()

        # Injection
        c = ws['C'+str(row)]
        c.value = risco
        c.font = Font(name='Segoe UI', size=9, bold=True, color=clr_r['fg'])
        c.fill = fill(clr_r['bg'])
        c.alignment = al('center')
        c.border = bdb(clr_r.get('bg','D0E4EE'))

        # Confiança
        c = ws['D'+str(row)]
        c.value = inj.get('confianca', '—')
        c.font = reg(9)
        c.alignment = al('center')
        c.border = bdb()

        # Nível
        c = ws['E'+str(row)]
        c.value = f'Nível {niv}' if niv else '—'
        c.font = Font(name='Segoe UI', size=9, bold=True, color='17324D')
        c.fill = fill(clr_n) if niv else fill(bg)
        c.alignment = al('center')
        c.border = bdb()

        # Valor
        c = ws['F'+str(row)]
        val = agr.get('valor_estimado', 0)
        c.value = f'R$ {val:,.2f}'.replace(',','X').replace('.',',').replace('X','.') if val else '—'
        c.font = reg(9)
        c.alignment = al('center')
        c.border = bdb()

        # Resumo
        c = ws['G'+str(row)]
        c.value = agr.get('resumo', ia.get('observacoes', '—'))
        c.font = Font(name='Segoe UI', size=9, color='5B7A8E', italic=True)
        c.alignment = al('left', 'center', wrap=True)
        c.border = bdb()

        row += 1

    # ── Aba 2: Injeções Detalhadas ──
    ws2 = wb.create_sheet("Injeções")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 22
    ws2.column_dimensions['F'].width = 45
    ws2.column_dimensions['G'].width = 60
    ws2.column_dimensions['H'].width = 3

    ws2.row_dimensions[1].height = 26
    ws2.merge_cells('B1:G1')
    ws2['B1'].value = 'Injeções Detectadas — Detalhamento'
    ws2['B1'].font = Font(name='Segoe UI', size=12, bold=True, color=WHITE)
    ws2['B1'].fill = fill(NAVY2)
    ws2['B1'].alignment = al('left', 'center')
    for col in 'ACDEFGH': ws2[col+'1'].fill = fill(NAVY2)

    ws2.row_dimensions[2].height = 4
    for col in 'ABCDEFGH': ws2[col+'2'].fill = fill(CYAN)

    ws2.row_dimensions[3].height = 26
    _hdr_row(ws2, 3, [
        ('B', 'Arquivo'), ('C', 'Risco'), ('D', 'Confiança'),
        ('E', 'Tipo'), ('F', 'Descrição'), ('G', 'Trecho Suspeito')
    ])

    row2 = 4
    injecoes = [r for r in resultados
                if r.get('risco_final', {}).get('risco') not in ('LIMPO', None)]

    # Ordena CRÍTICO > ALTO > MÉDIO
    ORDEM = {'CRÍTICO': 3, 'ALTO': 2, 'MÉDIO': 1, 'LIMPO': 0}
    injecoes.sort(key=lambda r: ORDEM.get(r.get('risco_final', {}).get('risco', ''), 0), reverse=True)

    for i, r in enumerate(injecoes):
        inj = r.get('risco_final', {})
        ia  = r.get('ia', {})
        achados_ia   = ia.get('injection', {}).get('achados', [])
        achados_stat = r.get('static', {}).get('achados', [])
        todos_achados = achados_ia + achados_stat

        if not todos_achados:
            todos_achados = [{'tipo': '—', 'risco': inj.get('risco','—'),
                              'confianca': inj.get('confianca','—'),
                              'descricao': '—', 'trecho': '—'}]

        for achado in todos_achados:
            ws2.row_dimensions[row2].height = 52
            bg = GRAY if i % 2 == 0 else WHITE
            risco = achado.get('risco', inj.get('risco', '—'))
            clr_r = {'CRÍTICO': C_CRITICO, 'ALTO': C_ALTO, 'MÉDIO': C_MEDIO, 'LIMPO': C_LIMPO}.get(risco, C_LIMPO)

            for col in 'ABCDEFGH': ws2[col+str(row2)].fill = fill(bg)

            ws2['B'+str(row2)].value = r.get('filename', '')
            ws2['B'+str(row2)].font = bold(9)
            ws2['B'+str(row2)].alignment = al('left', 'center', wrap=True)
            ws2['B'+str(row2)].border = bdb()

            c = ws2['C'+str(row2)]
            c.value = risco
            c.font = Font(name='Segoe UI', size=9, bold=True, color=clr_r['fg'])
            c.fill = fill(clr_r['bg'])
            c.alignment = al('center')

            ws2['D'+str(row2)].value = achado.get('confianca', '—')
            ws2['D'+str(row2)].font = reg(9)
            ws2['D'+str(row2)].alignment = al('center')
            ws2['D'+str(row2)].border = bdb()

            ws2['E'+str(row2)].value = achado.get('tipo', '—')
            ws2['E'+str(row2)].font = reg(9)
            ws2['E'+str(row2)].alignment = al('left', 'center', wrap=True)
            ws2['E'+str(row2)].border = bdb()

            ws2['F'+str(row2)].value = achado.get('descricao', '—')
            ws2['F'+str(row2)].font = reg(9, '5B7A8E')
            ws2['F'+str(row2)].alignment = al('left', 'center', wrap=True)
            ws2['F'+str(row2)].border = bdb()

            ws2['G'+str(row2)].value = str(achado.get('trecho', '—'))[:500]
            ws2['G'+str(row2)].font = Font(name='Courier New', size=8, color='991B1B')
            ws2['G'+str(row2)].alignment = al('left', 'top', wrap=True)
            ws2['G'+str(row2)].border = bdb()

            row2 += 1

    # ── Aba 3: Análise de Risco ──
    ws3 = wb.create_sheet("Risco Jurídico")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 3
    ws3.column_dimensions['B'].width = 32
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 16
    ws3.column_dimensions['E'].width = 42
    ws3.column_dimensions['F'].width = 42
    ws3.column_dimensions['G'].width = 18
    ws3.column_dimensions['H'].width = 3

    ws3.row_dimensions[1].height = 26
    ws3.merge_cells('B1:G1')
    ws3['B1'].value = 'Análise de Risco Jurídico — Agressividade'
    ws3['B1'].font = Font(name='Segoe UI', size=12, bold=True, color=WHITE)
    ws3['B1'].fill = fill(NAVY2)
    ws3['B1'].alignment = al('left', 'center')
    for col in 'ACDEFGH': ws3[col+'1'].fill = fill(NAVY2)

    ws3.row_dimensions[2].height = 4
    for col in 'ABCDEFGH': ws3[col+'2'].fill = fill(CYAN)

    ws3.row_dimensions[3].height = 26
    _hdr_row(ws3, 3, [
        ('B', 'Arquivo'), ('C', 'Nível'), ('D', 'Valor Estimado'),
        ('E', 'Verbas / Teses'), ('F', 'Pontos de Atenção'), ('G', 'Confiança IA')
    ])

    row3 = 4
    sorted_r = sorted(resultados,
                      key=lambda r: r.get('ia',{}).get('agressividade',{}).get('nivel',0),
                      reverse=True)

    for i, r in enumerate(sorted_r):
        ws3.row_dimensions[row3].height = 52
        bg = GRAY if i % 2 == 0 else WHITE
        ia  = r.get('ia', {})
        agr = ia.get('agressividade', {})
        niv = agr.get('nivel', 0)
        clr_n = C_NIV.get(niv, WHITE)

        for col in 'ABCDEFGH': ws3[col+str(row3)].fill = fill(bg)

        ws3['B'+str(row3)].value = r.get('filename','')
        ws3['B'+str(row3)].font = bold(9)
        ws3['B'+str(row3)].alignment = al('left','center',wrap=True)
        ws3['B'+str(row3)].border = bdb()

        ws3['C'+str(row3)].value = f'N{niv}' if niv else '—'
        ws3['C'+str(row3)].font = Font(name='Segoe UI',size=10,bold=True,color='17324D')
        ws3['C'+str(row3)].fill = fill(clr_n) if niv else fill(bg)
        ws3['C'+str(row3)].alignment = al('center')
        ws3['C'+str(row3)].border = bdb()

        val = agr.get('valor_estimado',0)
        ws3['D'+str(row3)].value = f'R$ {val:,.2f}'.replace(',','X').replace('.',',').replace('X','.') if val else '—'
        ws3['D'+str(row3)].font = reg(9)
        ws3['D'+str(row3)].alignment = al('center')
        ws3['D'+str(row3)].border = bdb()

        verbas = agr.get('verbas_pedidas',[])
        teses  = agr.get('teses_juridicas',[])
        vt = '• ' + '\n• '.join(verbas[:5]) if verbas else '—'
        if teses: vt += '\n\nTeses: ' + '; '.join(teses[:3])
        ws3['E'+str(row3)].value = vt
        ws3['E'+str(row3)].font = reg(9,'5B7A8E')
        ws3['E'+str(row3)].alignment = al('left','top',wrap=True)
        ws3['E'+str(row3)].border = bdb()

        pontos = agr.get('pontos_atencao',[])
        ws3['F'+str(row3)].value = ('• ' + '\n• '.join(pontos[:4])) if pontos else '—'
        ws3['F'+str(row3)].font = reg(9,'991B1B')
        ws3['F'+str(row3)].alignment = al('left','top',wrap=True)
        ws3['F'+str(row3)].border = bdb()

        ws3['G'+str(row3)].value = agr.get('confianca','—')
        ws3['G'+str(row3)].font = reg(9)
        ws3['G'+str(row3)].alignment = al('center')
        ws3['G'+str(row3)].border = bdb()

        row3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
